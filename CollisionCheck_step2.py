import pip._internal as pip
# Импорт библиотеки openpyxl
def import_lib(name):
    try:
        return __import__(name) # пытаемся импортировать
    except ImportError:
        pip.main(['install', name]) # ставим библиотеку если её нет
    return __import__(name) # возвращаем библиотеку

if __name__ == '__main__':
    numpy = import_lib('openpyxl')
    import openpyxl

    book = openpyxl.open('test.xlsx')  # указать имя исходног файла excel


    s1 = book['Модели']  # лист с списком моделей
    s2 = book['Матрица']  # лист с матрицей коллизий
    s3 = book['Модели_Элементы']  # лист с элементами иоделей, созданный на первом шаге

    # удаление и создание новых листов
    if 'Матрица_Заполнение' in book.sheetnames:
        del book['Матрица_Заполнение']
    if 'Списки_Заполнение' in book.sheetnames:
        del book['Списки_Заполнение']
    s4 = book.create_sheet('Матрица_Заполнение')
    s5 = book.create_sheet('Списки_Заполнение')

    models = list(s1.values)  # список моделей по разделам
    matrix_0 = list(s2.values)  # исходная матрица коллизий

    selectionsets_list = list(s3.values)

    matrix_2 = [[''] * len(selectionsets_list) for i in range(len(selectionsets_list))]  # пустая матрица необходимого размера
    clashtests = []
    selectionsets = [[], []]
    tolerance = []
    clashtests_models = [[], []]
    # заполенние пустой матрицы названиями проверок
    x = -1
    for i in range(len(selectionsets_list)):  # строки
        x += 1
        for j in range(0 + x, len(selectionsets_list)):  # столбцы
            for k in range(2, len(matrix_0)):  # строки
                for n in range(2, len(matrix_0[k])):  # столбцы
                    if (selectionsets_list[j][0] == matrix_0[0][n] and selectionsets_list[j][3] == matrix_0[1][n]
                            and selectionsets_list[i][0] == matrix_0[k][0] and selectionsets_list[i][3] == matrix_0[k][1]):
                        matrix_2[i][j] = (selectionsets_list[i][0] + ' vs ' + selectionsets_list[j][0] + '-' + selectionsets_list[i][2] +
                                          ' vs ' + selectionsets_list[j][2] + '-' + selectionsets_list[i][3] + ' vs '
                                          + selectionsets_list[j][3] + '-' + str(matrix_0[k][n]))
                        clashtests.append(selectionsets_list[i][0] + ' vs ' + selectionsets_list[j][0] + '-' + selectionsets_list[i][2]
                                          + ' vs ' + selectionsets_list[j][2] + '-' + selectionsets_list[i][3] + ' vs ' + selectionsets_list[j][3])
                        selectionsets[0].append(selectionsets_list[i][1])
                        selectionsets[1].append(selectionsets_list[j][1])
                        tolerance.append(str(matrix_0[k][n]))
                        clashtests_models[0].append(selectionsets_list[i][2])
                        clashtests_models[1].append(selectionsets_list[j][2])

    # заголовки матрицы коллизий
    for i in range(len(selectionsets_list)):
        s4.cell(row=1, column=i + 4, value=selectionsets_list[i][0])
        s4.cell(row=2, column=i + 4, value=selectionsets_list[i][2])
        s4.cell(row=3, column=i + 4, value=selectionsets_list[i][3])
        s4.cell(row=i + 4, column=1, value=selectionsets_list[i][0])
        s4.cell(row=i + 4, column=2, value=selectionsets_list[i][2])
        s4.cell(row=i + 4, column=3, value=selectionsets_list[i][3])

    # наполнение матрицы коллищий
    for i in range(len(matrix_2)):  # строка
        for j in range(len(matrix_2[i])):  # столбец
            s4.cell(row=i + 4, column=j + 4, value=matrix_2[i][j])

    for i in range(len(clashtests)):
        s5.cell(row=i + 1, column=1, value=clashtests[i])
    for i in range(len(selectionsets[0])):
        s5.cell(row=i + 1, column=2, value=selectionsets[0][i])
        s5.cell(row=i + 1, column=3, value=selectionsets[1][i])
    for i in range(len(tolerance)):
        s5.cell(row=i + 1, column=4, value=tolerance[i])

    # шаблон для файла xml с првоерками
    xml_template_clashtests = '''<?xml version="1.0" encoding="UTF-8"?>  
    
<exchange xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://download.autodesk.com/us/navisworks/schemas/nw-exchange-12.0.xsd" units="m" filename="Проверки" filepath="">  
    <batchtest name="Проверки" internal_name="Проверки" units="m">    
        <clashtests>      
            {0}
        </clashtests>    
        <selectionsets>    
        </selectionsets>  
    </batchtest>
</exchange>'''

    xml_template_clashtest = '''<clashtest name="{0}" test_type="hard" status="new" tolerance="{1}" merge_composites="1">  
                <linkage mode="none"/>        
                <left>          
                    <clashselection selfintersect="0" primtypes="1">            
                        <locator>lcop_selection_set_tree/Поисковые наборы по моделям/{2}/{3}</locator>          
                    </clashselection>        
                </left>       
                <right>          
                    <clashselection selfintersect="0" primtypes="1">            
                        <locator>lcop_selection_set_tree/Поисковые наборы по моделям/{4}/{5}</locator>          
                    </clashselection>        
                </right>        
                <rules/>      
            </clashtest>
            '''
    xml_clashtests_str = ''

    for i, j, k, n, h, t in zip(clashtests, selectionsets[0], selectionsets[1], tolerance, clashtests_models[0], clashtests_models[1]):
        xml_clashtests_str += xml_template_clashtest.format(i, n, h, j, t, k)

    xml_clashtests = xml_template_clashtests.format(xml_clashtests_str)
    xml_doc_clashtests = open('Проверки.xml', 'w', encoding='utf-8')  # при создании файла указал кодировку, чтобы кириллица записалась корректно
    xml_doc_clashtests.write(xml_clashtests)

    # шаблон для файла xml с поисковыми наборами
    xml_template_selectionsets = '''<?xml version="1.0" encoding="UTF-8"?>  
    
<exchange xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://download.autodesk.com/us/navisworks/schemas/nw-exchange-12.0.xsd" units="ft" filename="" filepath="">  
    <selectionsets>
        <viewfolder name="Модели" guid="">
            {0}
        </viewfolder>
        <viewfolder name="Поисковые наборы по моделям" guid="">
            {1}
        </viewfolder>
    </selectionsets>
</exchange>'''

    xml_template_selectionset_folder = '''<viewfolder name="{0}" guid="">
                {1}
            </viewfolder>
            '''

    xml_template_selectionset = '''<selectionset name="{0}" guid="">  
                    <findspec mode="all" disjoint="0">        
                        <conditions>     
                            <condition test="contains" flags="10">            
                                <property>              
                                    <name internal="LcOaNodeSourceFile">Файл источника</name>            
                                </property>            
                                <value>              
                                    <data type="wstring">{1}</data>            
                                </value>          
                            </condition>                 
                            <condition test="contains" flags="10">            
                                <category>              
                                    <name internal="LcRevitData_Element">Объект</name>            
                                </category>            
                                <property>              
                                    <name internal="LcRevitPropertyElementCategory">Категория</name>            
                                </property>            
                                <value>              
                                    <data type="wstring">{2}</data>            
                                </value>          
                            </condition>        
                        </conditions>        
                        <locator>/</locator>      
                    </findspec>    
                </selectionset>
                '''
    xml_template_selectionset_models = '''<selectionset name="{0}" guid="">  
                <findspec mode="all" disjoint="0">        
                    <conditions>          
                        <condition test="contains" flags="10">            
                            <property>              
                                <name internal="LcOaNodeSourceFile">Файл источника</name>            
                            </property>            
                            <value>              
                                <data type="wstring">{0}</data>            
                            </value>          
                        </condition>                
                    </conditions>        
                    <locator>/</locator>      
                </findspec>    
            </selectionset>
            '''

    xml_selectionset_list = []
    for model in models:
        xml_selectionsets_str = ''
        for i in selectionsets_list:
            if i[2] == model[0]:
                xml_selectionsets_str += xml_template_selectionset.format(i[1], i[2], i[3])
        xml_selectionset_list.append(xml_selectionsets_str)

    xml_selectionset_folder_str = ''
    for model, folder in zip(models, xml_selectionset_list):
        xml_selectionset_folder_str += xml_template_selectionset_folder.format(model[0], folder)

    xml_selectionset_models_str = ''
    for model in models:
        xml_selectionset_models_str += xml_template_selectionset_models.format(model[0])

    xml_selectionsets = xml_template_selectionsets.format(xml_selectionset_models_str, xml_selectionset_folder_str)

    xml_doc_selectionsets = open('Поисковые наборы.xml', 'w', encoding='utf-8')  # при создании файла указал кодировку, чтобы кириллица записалась корректно
    xml_doc_selectionsets.write(xml_selectionsets)

    book.save("test.xlsx")
