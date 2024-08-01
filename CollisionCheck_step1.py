import pip._internal as pip

def import_lib(name):
    try:
        return __import__(name) # пытаемся импортировать
    except ImportError:
        pip.main(['install', name]) # ставим библиотеку если её нет
    return __import__(name) # возвращаем библиотеку

if __name__ == '__main__':
    numpy = import_lib('openpyxl')
    import openpyxl

    book = openpyxl.open('test.xlsx')  # указать имя исходног файла excel


    s1 = book['Модели']  # лист с списком моделей
    s2 = book['Матрица']  # лист с матрицей коллизий

    # удаление и создание новых листов
    if 'Модели_Элементы' in book.sheetnames:
        del book['Модели_Элементы']
    s3 = book.create_sheet('Модели_Элементы')

    models = list(s1.values)  # список моделей по разделам
    matrix_0 = list(s2.values)  # исходная матрица коллизий

    selectionsets_list = []
    for i in models:
        for j in matrix_0:
            if i[1] == j[0]:
                selectionsets_list.append([i[1], i[0] + '-' + j[1], i[0], j[1]])

    for i in range(len(selectionsets_list)):
        s3.cell(row=i + 1, column=1, value=selectionsets_list[i][0])
        s3.cell(row=i + 1, column=2, value=selectionsets_list[i][1])
        s3.cell(row=i + 1, column=3, value=selectionsets_list[i][2])
        s3.cell(row=i + 1, column=4, value=selectionsets_list[i][3])

    book.save("test.xlsx")
