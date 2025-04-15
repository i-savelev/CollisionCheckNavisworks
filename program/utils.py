import openpyxl
import xml_template

def step_1(path_excel_file):

    book = openpyxl.open(path_excel_file)

    s1 = book['Модели']  
    s2 = book['Матрица']  
    
    if 'Модели_Элементы' in book.sheetnames:
        del book['Модели_Элементы']
    s3 = book.create_sheet('Модели_Элементы')

    models = list(s1.values)  
    matrix_0 = list(s2.values) 

    selectionsets_list = []
    for i in models:
        for j in matrix_0:
            if i[1] == j[0]:
                selectionsets_list.append([i[1], i[0] + '-' + j[1], i[0], j[1]])

    for i in range(len(selectionsets_list)):
        s3.cell(row=i + 1, column=1, value=selectionsets_list[i][0])
        s3.cell(row=i + 1, column=2, value=selectionsets_list[i][1])
        s3.cell(row=i + 1, column=3, value=selectionsets_list[i][2])
        s3.cell(row=i + 1, column=4, value=selectionsets_list[i][3])

    book.save(path_excel_file)

def step_2(path_excel_file, folder_xml_file):

    book = openpyxl.open(path_excel_file)  # указать имя исходног файла excel
    s1 = book['Модели'] 
    s2 = book['Матрица'] 
    s3 = book['Модели_Элементы']  

    if 'Матрица_Заполнение' in book.sheetnames:
        del book['Матрица_Заполнение']
    if 'Списки_Заполнение' in book.sheetnames:
        del book['Списки_Заполнение']
    s4 = book.create_sheet('Матрица_Заполнение')
    s5 = book.create_sheet('Списки_Заполнение')

    models = list(s1.values)  # список моделей по разделам
    matrix_0 = list(s2.values)  # исходная матрица коллизий

    selectionsets_list = list(s3.values)

    matrix_2 = [[''] * len(selectionsets_list) for i in range(len(selectionsets_list))]  # пустая матрица необходимого размера
    clashtests = []
    selectionsets = [[], []]
    tolerance = []
    clashtests_models = [[], []]
    # заполенние пустой матрицы названиями проверок
    x = -1
    for i in range(len(selectionsets_list)):  # строки
        x += 1
        for j in range(0 + x, len(selectionsets_list)):  # столбцы
            for k in range(2, len(matrix_0)):  # строки
                for n in range(2, len(matrix_0[k])):  # столбцы
                    if (selectionsets_list[j][0] == matrix_0[0][n] and selectionsets_list[j][3] == matrix_0[1][n]
                            and selectionsets_list[i][0] == matrix_0[k][0] and selectionsets_list[i][3] == matrix_0[k][1]):
                        matrix_2[i][j] = (selectionsets_list[i][0] + ' vs ' + selectionsets_list[j][0] + '-' + selectionsets_list[i][2] +
                                          ' vs ' + selectionsets_list[j][2] + '-' + selectionsets_list[i][3] + ' vs '
                                          + selectionsets_list[j][3] + '-' + str(matrix_0[k][n]))
                        clashtests.append(selectionsets_list[i][0] + ' vs ' + selectionsets_list[j][0] + '-' + selectionsets_list[i][2]
                                          + ' vs ' + selectionsets_list[j][2] + '-' + selectionsets_list[i][3] + ' vs ' + selectionsets_list[j][3])
                        selectionsets[0].append(selectionsets_list[i][1])
                        selectionsets[1].append(selectionsets_list[j][1])
                        tolerance.append(str(matrix_0[k][n]))
                        clashtests_models[0].append(selectionsets_list[i][2])
                        clashtests_models[1].append(selectionsets_list[j][2])

    # заголовки матрицы коллизий
    for i in range(len(selectionsets_list)):
        s4.cell(row=1, column=i + 4, value=selectionsets_list[i][0])
        s4.cell(row=2, column=i + 4, value=selectionsets_list[i][2])
        s4.cell(row=3, column=i + 4, value=selectionsets_list[i][3])
        s4.cell(row=i + 4, column=1, value=selectionsets_list[i][0])
        s4.cell(row=i + 4, column=2, value=selectionsets_list[i][2])
        s4.cell(row=i + 4, column=3, value=selectionsets_list[i][3])

    # наполнение матрицы коллизий
    for i in range(len(matrix_2)):  # строка
        for j in range(len(matrix_2[i])):  # столбец
            s4.cell(row=i + 4, column=j + 4, value=matrix_2[i][j])

    for i in range(len(clashtests)):
        s5.cell(row=i + 1, column=1, value=clashtests[i])
    for i in range(len(selectionsets[0])):
        s5.cell(row=i + 1, column=2, value=selectionsets[0][i])
        s5.cell(row=i + 1, column=3, value=selectionsets[1][i])
    for i in range(len(tolerance)):
        s5.cell(row=i + 1, column=4, value=tolerance[i])

    xml_clashtests_str = ''

    for i, j, k, n, h, t in zip(clashtests, selectionsets[0], selectionsets[1], tolerance, clashtests_models[0], clashtests_models[1]):
        xml_clashtests_str += xml_template.xml_template_clashtest.format(i, n, h, j, t, k)

    xml_clashtests = xml_template.xml_template_clashtests.format(xml_clashtests_str)
    xml_doc_clashtests = open(f'{folder_xml_file}/Проверки.xml', 'w', encoding='utf-8')  
    xml_doc_clashtests.write(xml_clashtests)

    xml_selectionset_list = []
    for model in models:
        xml_selectionsets_str = ''
        for i in selectionsets_list:
            if i[2] == model[0]:
                xml_selectionsets_str += xml_template.xml_template_selectionset.format(i[1], i[2], i[3])
        xml_selectionset_list.append(xml_selectionsets_str)

    xml_selectionset_folder_str = ''
    for model, folder in zip(models, xml_selectionset_list):
        xml_selectionset_folder_str += xml_template.xml_template_selectionset_folder.format(model[0], folder)

    xml_selectionset_models_str = ''
    for model in models:
        xml_selectionset_models_str += xml_template.xml_template_selectionset_models.format(model[0])

    xml_selectionsets = xml_template.xml_template_selectionsets.format(xml_selectionset_models_str, xml_selectionset_folder_str)
    xml_doc_selectionsets = open(f'{folder_xml_file}/Поисковые наборы.xml', 'w', encoding='utf-8')
    xml_doc_selectionsets.write(xml_selectionsets)

    book.save(path_excel_file)